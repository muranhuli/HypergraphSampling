import networkx as nx
from algorithms import *
from tools import *
from collections import Counter
from operator import itemgetter
import scipy.stats
import numpy as np
import math
import sys
import openpyxl
import pandas as pd
import os


filepath = './datasets/standardized_hypergraph/'
file_list = 'NDCC TaMS CoBi TaAU MaAn WaTr ThAU ThMS TrCl'.split()

for algorithm_index in range(3):
    wb_node = openpyxl.Workbook()
    wb_edge = openpyxl.Workbook()
    for filename in file_list:
        print(algorithm_index, filename)
        graph_path = filepath + filename
        G0 = Hypergraph(graph_path)
        # Select a function to sample
        if algorithm_index == 0:
            S = RandomWalkSampler(int(G0.number_of_nodes()*0.5))
        elif algorithm_index ==1:
            S = RandomWalkSamplerWithNonBacktracking(int(G0.number_of_nodes()*0.5))
        elif algorithm_index ==2:
            S = RandomWalkSamplerWithHistoryAware(int(G0.number_of_nodes()*0.5))
        # start sample
        sample_data = S.sample(G0)
        # Information on the node of the statistical sampling data
        SV = NodeStandardValue(G0,sample_data,0)
        ws_node = wb_node.create_sheet(filename)
        data = ['range', 'effective_sampling_node_size_ratio', 'total_variance_distance', 'KL', 'JS', 'relative_error_of_total_node_size', 'relative_error_of_average_degree', 'variance_of_sample_node', 'geweke_score']
        ws_node.append(data)
        for index in range(int(math.ceil(int(G0.number_of_nodes()*0.5) / 100))):
            SV = NodeStandardValue(G0,sample_data,index)
            effective_sampling_node_size_ratio = SV.effective_sampling_node_size_ratio()
            total_variance_distance = SV.total_variance_distance()
            KL, JS = SV.KL_JS()
            relative_error_of_total_node_size = SV.relative_error_of_total_node_size()
            relative_error_of_average_degree = SV.relative_error_of_average_degree()
            variance_of_sample_node = SV.variance_of_sample_node()
            geweke_score = SV.geweke_score()
            data = [f'(0,{(index+1)*100})', effective_sampling_node_size_ratio, total_variance_distance, KL, JS, relative_error_of_total_node_size, relative_error_of_average_degree, variance_of_sample_node, geweke_score]
            ws_node.append(data)

        # Information on the edge of the statistical sampling data
        SV = EdgeStandardValue(G0,sample_data,0)
        ws_edge = wb_edge.create_sheet(filename)
        data = ['range', 'effective_sampling_edge_size_ratio', 'total_variance_distance', 'KL', 'JS', 'relative_error_of_total_edge_size', 'relative_error_of_average_cardi', 'variance_of_sample_edge', 'geweke_score']
        ws_edge.append(data)
        for index in range(int(math.ceil(int(G0.number_of_nodes()*0.5) / 100))):
            SV = EdgeStandardValue(G0,sample_data,index)
            effective_sampling_edge_size_ratio = SV.effective_sampling_edge_size_ratio()
            total_variance_distance = SV.total_variance_distance()
            KL, JS = SV.KL_JS()
            relative_error_of_total_edge_size = SV.relative_error_of_total_edge_size()
            relative_error_of_average_cardi = SV.relative_error_of_average_cardi()
            variance_of_sample_edge = SV.variance_of_sample_edge()
            geweke_score = SV.geweke_score()
            data = [f'(0,{(index+1)*100})', effective_sampling_edge_size_ratio, total_variance_distance, KL, JS, relative_error_of_total_edge_size, relative_error_of_average_cardi, variance_of_sample_edge, geweke_score]
            ws_edge.append(data)
    
    # save data
    ws_node = wb_node['Sheet']
    wb_node.remove(ws_node)
    wb_node.save(f'./result/Node_{S.__class__.__name__}.xlsx')
    print(f'Node_{S.__class__.__name__}.xlsx has been generated!')        
    
    ws_edge = wb_edge['Sheet']
    wb_edge.remove(ws_edge)
    wb_edge.save(f'./result/Edge_{S.__class__.__name__}.xlsx')
    print(f'Edge_{S.__class__.__name__}.xlsx has been generated!')